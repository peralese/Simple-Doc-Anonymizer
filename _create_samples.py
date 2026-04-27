"""One-time script to generate sample input files."""
import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path

Path("input").mkdir(exist_ok=True)

wb = openpyxl.Workbook()

# ── Sheet 1: Contacts ──────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Contacts"

headers = ["Name", "Email", "Phone", "Department"]
ws1.append(headers)
for cell in ws1[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9E1F2")

contacts = [
    ["John Smith",      "john.smith@acme.com",    "+1-212-555-0141", "Engineering"],
    ["Sarah Mitchell",  "sarah.mitchell@acme.com", "+1-415-555-0192", "Product"],
    ["Linda Zhao",      "linda.zhao@acme.com",    "+1-650-555-0173", "Security"],
    ["Robert Huang",    "r.huang@acme.com",       "+1-312-555-0108", "Finance"],
    ["Alice Pemberley", "alice.pemberley@acme.com","+1-617-555-0155", "Legal"],
]
for row in contacts:
    ws1.append(row)

ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 30
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 16

# ── Sheet 2: Projects ──────────────────────────────────────────────────────
ws2 = wb.create_sheet("Projects")

proj_headers = ["ProjectName", "Lead", "Budget", "Notes"]
ws2.append(proj_headers)
for cell in ws2[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="E2EFDA")

projects = [
    [
        "Project Falcon",
        "Kevin Tran",
        "$1,250,000",
        "Prod deployment on PROD-DB-01. API key: sk-prod-aBcD1234eFgH5678iJkL",
    ],
    [
        "TIGERS Integration",
        "Sarah Mitchell",
        "$340,000",
        "eReporting module dependency. Contact sarah.mitchell@acme.com for access.",
    ],
    [
        "Compliance Refresh",
        "Alice Pemberley",
        "$95,000",
        "Initiated by Acme Corporation legal team. SSN audit complete.",
    ],
]
for row in projects:
    ws2.append(row)

ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 60

wb.save("input/sample_document.xlsx")
print("Created input/sample_document.xlsx")
